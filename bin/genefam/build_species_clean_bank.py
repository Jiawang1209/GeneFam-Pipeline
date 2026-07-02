#!/usr/bin/env python3
"""Build a reusable species clean bank from a raw folder-per-species bank."""

from __future__ import annotations

import argparse
import csv
import re
import shutil
import subprocess
import sys
from collections import Counter
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from bin.genefam.discover_species import FILE_TYPES, discover_species
from bin.genefam.preprocess_species import (
    ID_RULE_FIELDS,
    REPRESENTATIVE_FIELDS,
    TRANSCRIPT_FIELDS,
    WARNING_FIELDS,
    clean_header_id,
    clean_sequence_records,
    infer_species_id_rule,
    parse_gff3_transcript_gene_map,
    read_fasta_records,
    write_fasta,
    write_tsv,
)


DEFAULT_PATTERNS = {
    "pep": ["*.pep.fa", "*.pep.fasta", "*.protein.fa", "*.faa"],
    "cds": ["*.cds.fa", "*.cds.fasta"],
    "genome": ["*.genome.fa", "*.genome.fasta", "*.dna.fa", "*.dna.fasta", "*.fna"],
    "gff3": ["*.gff3", "*.gff"],
}
DEFAULT_REQUIRED = {"pep": True, "cds": False, "genome": False, "gff3": False}
MANIFEST_FIELDS = [
    "species_id",
    "protein",
    "cds",
    "genome",
    "gff3",
    "genome_lengths",
    "chromosome_lengths",
    "raw_pep",
    "raw_cds",
    "raw_genome",
    "raw_gff3",
    "gene_id_map",
    "id_resolution_rules",
    "representative_transcripts",
    "preprocess_qc",
    "preprocess_warnings",
    "status",
]
QC_FIELDS = [
    "species_id",
    "raw_pep_count",
    "raw_cds_count",
    "clean_pep_count",
    "clean_cds_count",
    "genome_seq_count",
    "chromosome_seq_count",
    "unassembled_seq_count",
    "organelle_seq_count",
    "total_genome_bp",
    "chromosome_bp",
    "assembly_level",
    "chromosome_analysis_ready",
    "gff3_transcript_map_count",
    "gff3_mapping_rate",
    "fallback_mapping_rate",
    "cds_match_rate",
    "terminal_stop_removed_count",
    "warning_count",
    "status",
    "note",
]
FAILED_FIELDS = ["species_id", "status", "reason"]
GENOME_LENGTH_FIELDS = ["SeqID", "Start", "End", "Length", "SeqType", "Header"]
CHROMOSOME_LENGTH_FIELDS = ["Chr", "Start", "End"]
SPECIES_INFO_FIELDS = ["species_id", "latin_name"]
SPECIES_TREE_STATUS_FIELDS = ["source", "status", "species_count", "tree", "note"]


def iter_fasta_lengths(path: Path | str) -> list[tuple[str, str, int]]:
    lengths: list[tuple[str, str, int]] = []
    current_header: str | None = None
    current_length = 0
    with Path(path).open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if current_header is not None:
                    seqid = clean_header_id(current_header)
                    lengths.append((seqid, current_header, current_length))
                current_header = line[1:]
                current_length = 0
            elif current_header is None:
                raise ValueError(f"Sequence found before first FASTA header in {path}")
            else:
                current_length += len(line)
    if current_header is not None:
        seqid = clean_header_id(current_header)
        lengths.append((seqid, current_header, current_length))
    return lengths


def classify_genome_sequence(seqid: str, header: str) -> str:
    value = f"{seqid} {header}".casefold()
    if re.search(r"(^|[^a-z0-9])(chrm|chrmt|mt|mitochond)", value):
        return "organelle"
    if re.search(r"(^|[^a-z0-9])(chrc|chrcp|cp|chloroplast|plastid)", value):
        return "organelle"
    if re.search(r"(scaffold|contig|unplaced|unlocalized|random|chrur|chrun|chrsy|^nw_|^nt_)", value):
        return "unassembled"
    if re.match(r"^(chr)?[0-9]+[a-z]?$", seqid, flags=re.IGNORECASE):
        return "chromosome"
    if re.match(r"^chr[0-9][0-9]?$", seqid, flags=re.IGNORECASE):
        return "chromosome"
    if re.match(r"^[A-Z][a-z]{1,4}[0-9]{1,2}[A-Z]?$", seqid):
        return "chromosome"
    if re.search(r"(^|[^a-z0-9])chromosome([^a-z0-9]|$)", value):
        return "chromosome"
    return "unclassified"


def should_promote_scaffold_to_chromosome(seqid: str, length: int, scaffold_chromosome_min_bp: int) -> bool:
    if scaffold_chromosome_min_bp <= 0 or length < scaffold_chromosome_min_bp:
        return False
    return re.match(r"^scaffold[_-]?[0-9]+$", seqid, flags=re.IGNORECASE) is not None


def build_genome_length_rows(
    path: Path | str | None,
    scaffold_chromosome_min_bp: int = 0,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    if not path:
        return [], []
    genome_rows: list[dict[str, str]] = []
    chromosome_rows: list[dict[str, str]] = []
    for seqid, header, length in iter_fasta_lengths(path):
        seq_type = classify_genome_sequence(seqid, header)
        if seq_type == "unassembled" and should_promote_scaffold_to_chromosome(
            seqid,
            length,
            scaffold_chromosome_min_bp,
        ):
            seq_type = "chromosome"
        row = {
            "SeqID": seqid,
            "Start": "1",
            "End": str(length),
            "Length": str(length),
            "SeqType": seq_type,
            "Header": header,
        }
        genome_rows.append(row)
        if seq_type == "chromosome":
            chromosome_rows.append({"Chr": seqid, "Start": "1", "End": str(length)})
    return genome_rows, chromosome_rows


def count_fasta_records(path: Path | str | None) -> int:
    if not path:
        return 0
    return len(read_fasta_records(Path(path)))


def terminal_stop_count(records: list[tuple[str, str]]) -> int:
    return sum(1 for _record_id, sequence in records if sequence.endswith("*"))


def write_rows(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)


def latin_name(species_id: str) -> str:
    return species_id.replace("_", " ")


def successful_species_rows(qc_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return [
        {"species_id": row["species_id"], "latin_name": latin_name(row["species_id"])}
        for row in sorted(qc_rows, key=lambda item: item["species_id"])
        if row.get("status") in {"pass", "warning"}
    ]


def write_species_info(txt_path: Path, tsv_path: Path, qc_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    rows = successful_species_rows(qc_rows)
    txt_path.parent.mkdir(parents=True, exist_ok=True)
    txt_path.write_text("".join(f"{row['latin_name']}\n" for row in rows), encoding="utf-8")
    write_rows(tsv_path, SPECIES_INFO_FIELDS, rows)
    return rows


def write_species_tree_status(path: Path, row: dict[str, str]) -> None:
    write_rows(path, SPECIES_TREE_STATUS_FIELDS, [row])


def prepare_species_tree_outputs(
    *,
    source: str,
    user_tree: Path | None,
    species_rows: list[dict[str, str]],
    tree_dir: Path,
    timetree_run: bool = False,
) -> None:
    tree_dir.mkdir(parents=True, exist_ok=True)
    status_path = tree_dir / "species_tree_status.tsv"
    species_count = str(len(species_rows))
    if source == "none":
        return
    if source == "user":
        if user_tree is None or not user_tree.exists():
            write_species_tree_status(
                status_path,
                {
                    "source": "user",
                    "status": "missing_input",
                    "species_count": species_count,
                    "tree": "",
                    "note": "User species tree was not provided or does not exist",
                },
            )
            return
        destination = tree_dir / "species_tree.nwk"
        shutil.copy2(user_tree, destination)
        write_species_tree_status(
            status_path,
            {
                "source": "user",
                "status": "available",
                "species_count": species_count,
                "tree": str(destination.resolve()),
                "note": "Copied user-provided Newick tree",
            },
        )
        return
    if source == "timetree":
        species_input = tree_dir / "timetree_species_input.txt"
        species_input.write_text("".join(f"{row['latin_name']}\n" for row in species_rows), encoding="utf-8")
        if not timetree_run:
            write_species_tree_status(
                status_path,
                {
                    "source": "timetree",
                    "status": "pending_manual_upload",
                    "species_count": species_count,
                    "tree": "",
                    "note": "TimeTree input file is ready. Upload timetree_species_input.txt to https://timetree.org/ manually, or rerun with --species-tree-timetree-run to try browser automation.",
                },
            )
            return
        script = REPO_ROOT / "bin/genefam/fetch_timetree_species_tree.py"
        tree_path = tree_dir / "species_tree.nwk"
        if script.exists():
            completed = subprocess.run(
                [
                    sys.executable,
                    str(script),
                    "--species-list",
                    str(species_input),
                    "--out-tree",
                    str(tree_path),
                    "--status",
                    str(status_path),
                ],
                check=False,
                text=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
            )
            if status_path.exists():
                return
            note = completed.stdout.strip()[:500] if completed.stdout else "TimeTree automation did not write a status file"
        else:
            note = "TimeTree automation helper is not available"
        write_species_tree_status(
            status_path,
            {
                "source": "timetree",
                "status": "pending_manual_upload",
                "species_count": species_count,
                "tree": "",
                "note": f"Upload timetree_species_input.txt to https://timetree.org/ manually. {note}",
            },
        )
        return
    raise ValueError(f"Unsupported species tree source: {source}")


def write_qc_excel(path: Path, qc_rows: list[dict[str, str]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - depends on local runtime
        raise RuntimeError("openpyxl is required to write clean-bank Excel QC files") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "species_qc"
    worksheet.append(QC_FIELDS)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in qc_rows:
        worksheet.append([row.get(field, "") for field in QC_FIELDS])
    ready_col = QC_FIELDS.index("chromosome_analysis_ready") + 1
    for row_index in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_index, column=ready_col)
        if str(cell.value).upper() == "TRUE":
            cell.fill = PatternFill("solid", fgColor="C6EFCE")
        else:
            cell.fill = PatternFill("solid", fgColor="FFC7CE")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_index, field in enumerate(QC_FIELDS, start=1):
        values = [field] + [row.get(field, "") for row in qc_rows]
        width = min(max(len(str(value)) for value in values) + 2, 48)
        worksheet.column_dimensions[get_column_letter(col_index)].width = width

    summary = workbook.create_sheet("assembly_summary")
    summary.append(["metric", "value"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    ready_count = sum(1 for row in qc_rows if row.get("chromosome_analysis_ready") == "TRUE")
    assembly_counts = Counter(row.get("assembly_level", "unknown") for row in qc_rows)
    summary.append(["species_count", len(qc_rows)])
    summary.append(["chromosome_analysis_ready", ready_count])
    for level, count in sorted(assembly_counts.items()):
        summary.append([f"assembly_level:{level}", count])
    summary.column_dimensions["A"].width = 36
    summary.column_dimensions["B"].width = 18
    workbook.save(path)


def copy_if_present(source: str, destination: Path) -> str:
    if not source:
        return ""
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    return str(destination.resolve())


def symlink_if_present(source: str, destination: Path) -> str:
    return materialize_large_file(source, destination, "symlink")


def materialize_large_file(source: str, destination: Path, mode: str) -> str:
    if not source:
        return ""
    source_path = Path(source).resolve()
    if mode == "skip":
        return str(source_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists() or destination.is_symlink():
        destination.unlink()
    if mode == "copy":
        shutil.copy2(source_path, destination)
    elif mode == "symlink":
        destination.symlink_to(source_path)
    else:
        raise ValueError(f"Unsupported large-file mode: {mode}")
    return str(destination.absolute())


def _rate(numerator: int, denominator: int) -> str:
    if denominator == 0:
        return "0.0000"
    return f"{numerator / denominator:.4f}"


def _status(cleaned_pep_count: int, cleaned_cds_count: int, cds_required: bool, warning_count: int) -> tuple[str, str]:
    if cleaned_pep_count == 0:
        return "failed", "no clean protein records were produced"
    if cds_required and cleaned_cds_count == 0:
        return "missing_cds", "CDS is required but no clean CDS records were produced"
    if cleaned_cds_count not in {0, cleaned_pep_count}:
        return "warning", "clean protein and CDS counts differ"
    if warning_count:
        return "warning", "non-fatal preprocessing warnings were produced"
    return "pass", "clean bank species completed"


def infer_assembly_level(genome_seq_count: int, chromosome_seq_count: int) -> tuple[str, str]:
    if genome_seq_count == 0:
        return "missing_genome", "FALSE"
    if chromosome_seq_count > 0:
        return "chromosome", "TRUE"
    return "scaffold_or_contig", "FALSE"


def missing_input_reason(row: dict[str, str], *, cds_required: bool, genome_required: bool, gff3_required: bool) -> str:
    required = {"pep": True, "cds": cds_required, "genome": genome_required, "gff3": gff3_required}
    missing = [file_type for file_type, is_required in required.items() if is_required and not row.get(file_type)]
    if not missing:
        return ""
    return "missing required input file(s): " + ", ".join(missing)


def failed_qc_row(species_id: str, status: str, note: str) -> dict[str, str]:
    return {
        "species_id": species_id,
        "raw_pep_count": "0",
        "raw_cds_count": "0",
        "clean_pep_count": "0",
        "clean_cds_count": "0",
        "genome_seq_count": "0",
        "chromosome_seq_count": "0",
        "unassembled_seq_count": "0",
        "organelle_seq_count": "0",
        "total_genome_bp": "0",
        "chromosome_bp": "0",
        "assembly_level": status,
        "chromosome_analysis_ready": "FALSE",
        "gff3_transcript_map_count": "0",
        "gff3_mapping_rate": "0.0000",
        "fallback_mapping_rate": "0.0000",
        "cds_match_rate": "0.0000",
        "terminal_stop_removed_count": "0",
        "warning_count": "1",
        "status": status,
        "note": note,
    }


def build_species_clean_bank(
    *,
    raw_root: Path,
    out_root: Path,
    include: str | list[str] | None = "all",
    exclude: list[str] | None = None,
    cds_required: bool = False,
    genome_required: bool = False,
    gff3_required: bool = False,
    large_file_mode: str = "symlink",
    scaffold_chromosome_min_bp: int = 0,
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    rows = discover_species(
        root=raw_root,
        include=include,
        exclude=exclude or [],
        patterns=DEFAULT_PATTERNS,
        required={file_type: False for file_type in FILE_TYPES},
        base_dir=None,
    )
    manifest_rows: list[dict[str, str]] = []
    qc_rows: list[dict[str, str]] = []
    failed_rows: list[dict[str, str]] = []

    for row in rows:
        species_id = row["species_id"]
        species_root = out_root / species_id
        raw_dir = species_root / "raw"
        clean_dir = species_root / "clean"
        audit_dir = species_root / "audit"
        missing_reason = missing_input_reason(
            row,
            cds_required=cds_required,
            genome_required=genome_required,
            gff3_required=gff3_required,
        )
        if missing_reason:
            status = "missing_required_input"
            failed_rows.append({"species_id": species_id, "status": status, "reason": missing_reason})
            qc_rows.append(failed_qc_row(species_id, status, missing_reason))
            manifest_rows.append(
                {
                    "species_id": species_id,
                    "protein": "",
                    "cds": "",
                    "genome": "",
                    "gff3": "",
                    "genome_lengths": "",
                    "chromosome_lengths": "",
                    "raw_pep": row.get("pep", ""),
                    "raw_cds": row.get("cds", ""),
                    "raw_genome": row.get("genome", ""),
                    "raw_gff3": row.get("gff3", ""),
                    "gene_id_map": "",
                    "id_resolution_rules": "",
                    "representative_transcripts": "",
                    "preprocess_qc": "",
                    "preprocess_warnings": "",
                    "status": status,
                }
            )
            continue
        try:
            raw_pep = symlink_if_present(row.get("pep", ""), raw_dir / f"{species_id}.pep.fa")
            raw_cds = symlink_if_present(row.get("cds", ""), raw_dir / f"{species_id}.cds.fa")
            raw_genome = symlink_if_present(row.get("genome", ""), raw_dir / f"{species_id}.genome.fa")
            raw_gff3 = symlink_if_present(row.get("gff3", ""), raw_dir / f"{species_id}.gff3")

            pep_records = read_fasta_records(Path(row["pep"]))
            cds_records = read_fasta_records(Path(row["cds"])) if row.get("cds") else []
            transcript_gene_map = parse_gff3_transcript_gene_map(row.get("gff3"))
            cleaned_pep, cleaned_cds, transcript_rows, representative_rows, warnings = clean_sequence_records(
                species_id=species_id,
                pep_records=pep_records,
                cds_records=cds_records,
                transcript_gene_map=transcript_gene_map,
            )
            id_rule = infer_species_id_rule(species_id, pep_records, transcript_gene_map)

            protein_clean = clean_dir / f"{species_id}.protein.clean.fa"
            cds_clean = clean_dir / f"{species_id}.cds.clean.fa"
            genome_clean = clean_dir / f"{species_id}.genome.fa"
            gff3_clean = clean_dir / f"{species_id}.gff3"
            genome_lengths = audit_dir / f"{species_id}.genome.lengths.tsv"
            chromosome_lengths = clean_dir / f"{species_id}.chromosome.lengths.tsv"
            gene_id_map = audit_dir / f"{species_id}.gene_id_map.tsv"
            id_resolution_rules = audit_dir / f"{species_id}.id_resolution_rules.tsv"
            representative_tsv = audit_dir / f"{species_id}.representative_transcripts.tsv"
            warnings_tsv = audit_dir / f"{species_id}.preprocess_warnings.tsv"
            qc_tsv = audit_dir / f"{species_id}.preprocess_qc.tsv"
            genome_clean_path = ""
            gff3_clean_path = ""

            write_fasta(cleaned_pep, protein_clean)
            if row.get("cds"):
                write_fasta(cleaned_cds, cds_clean)
            if row.get("genome"):
                genome_clean_path = materialize_large_file(row["genome"], genome_clean, large_file_mode)
            if row.get("gff3"):
                gff3_clean_path = materialize_large_file(row["gff3"], gff3_clean, large_file_mode)
            genome_length_rows, chromosome_length_rows = build_genome_length_rows(
                row.get("genome", ""),
                scaffold_chromosome_min_bp=scaffold_chromosome_min_bp,
            )
            write_rows(genome_lengths, GENOME_LENGTH_FIELDS, genome_length_rows)
            write_rows(chromosome_lengths, CHROMOSOME_LENGTH_FIELDS, chromosome_length_rows)
            write_tsv(transcript_rows, TRANSCRIPT_FIELDS, gene_id_map)
            write_tsv([id_rule], ID_RULE_FIELDS, id_resolution_rules)
            write_tsv(representative_rows, REPRESENTATIVE_FIELDS, representative_tsv)
            write_tsv(warnings, WARNING_FIELDS, warnings_tsv)

            source_counts = Counter(record["source"] for record in transcript_rows)
            genome_type_counts = Counter(record["SeqType"] for record in genome_length_rows)
            total_genome_bp = sum(int(record["Length"]) for record in genome_length_rows)
            chromosome_bp = sum(int(record["End"]) for record in chromosome_length_rows)
            assembly_level, chromosome_analysis_ready = infer_assembly_level(
                len(genome_length_rows),
                len(chromosome_length_rows),
            )
            fallback_count = source_counts.get("auto_suffix", 0) + source_counts.get("self", 0)
            status, note = _status(len(cleaned_pep), len(cleaned_cds), cds_required, len(warnings))
            qc_row = {
                "species_id": species_id,
                "raw_pep_count": str(len(pep_records)),
                "raw_cds_count": str(len(cds_records)),
                "clean_pep_count": str(len(cleaned_pep)),
                "clean_cds_count": str(len(cleaned_cds)),
                "genome_seq_count": str(len(genome_length_rows)),
                "chromosome_seq_count": str(len(chromosome_length_rows)),
                "unassembled_seq_count": str(genome_type_counts.get("unassembled", 0)),
                "organelle_seq_count": str(genome_type_counts.get("organelle", 0)),
                "total_genome_bp": str(total_genome_bp),
                "chromosome_bp": str(chromosome_bp),
                "assembly_level": assembly_level,
                "chromosome_analysis_ready": chromosome_analysis_ready,
                "gff3_transcript_map_count": str(len(transcript_gene_map)),
                "gff3_mapping_rate": _rate(source_counts.get("gff3", 0), len(transcript_rows)),
                "fallback_mapping_rate": _rate(fallback_count, len(transcript_rows)),
                "cds_match_rate": _rate(len(cleaned_cds), len(cleaned_pep)),
                "terminal_stop_removed_count": str(terminal_stop_count(pep_records)),
                "warning_count": str(len(warnings)),
                "status": status,
                "note": note,
            }
            write_rows(qc_tsv, QC_FIELDS, [qc_row])
            qc_rows.append(qc_row)
            if status in {"failed", "missing_cds"}:
                failed_rows.append({"species_id": species_id, "status": status, "reason": note})
            manifest_rows.append(
                {
                    "species_id": species_id,
                    "protein": str(protein_clean.resolve()),
                    "cds": str(cds_clean.resolve()) if row.get("cds") else "",
                    "genome": genome_clean_path,
                    "gff3": gff3_clean_path,
                    "genome_lengths": str(genome_lengths.resolve()) if row.get("genome") else "",
                    "chromosome_lengths": str(chromosome_lengths.resolve()) if row.get("genome") else "",
                    "raw_pep": raw_pep,
                    "raw_cds": raw_cds,
                    "raw_genome": raw_genome,
                    "raw_gff3": raw_gff3,
                    "gene_id_map": str(gene_id_map.resolve()),
                    "id_resolution_rules": str(id_resolution_rules.resolve()),
                    "representative_transcripts": str(representative_tsv.resolve()),
                    "preprocess_qc": str(qc_tsv.resolve()),
                    "preprocess_warnings": str(warnings_tsv.resolve()),
                    "status": status,
                }
            )
        except Exception as exc:  # pragma: no cover - defensive per-species continuation
            reason = str(exc)
            failed_rows.append({"species_id": species_id, "status": "failed", "reason": reason})
            qc_rows.append(failed_qc_row(species_id, "failed", reason))
    return manifest_rows, qc_rows, failed_rows


def write_summary(path: Path, qc_rows: list[dict[str, str]], failed_rows: list[dict[str, str]]) -> None:
    status_counts = Counter(row["status"] for row in qc_rows)
    assembly_counts = Counter(row.get("assembly_level", "unknown") for row in qc_rows)
    ready_count = sum(1 for row in qc_rows if row.get("chromosome_analysis_ready") == "TRUE")
    missing_chromosome_rows = [
        row for row in qc_rows if row.get("genome_seq_count", "0") != "0" and row.get("chromosome_seq_count", "0") == "0"
    ]
    lines = [
        "# Species Clean Bank Summary",
        "",
        f"Species processed: {len(qc_rows)}",
        f"Passed: {status_counts.get('pass', 0)}",
        f"Warnings: {status_counts.get('warning', 0)}",
        f"Failed: {len(failed_rows)}",
        f"Chromosome-level analysis ready: {ready_count}",
        "",
        "| status | count |",
        "| --- | ---: |",
    ]
    for status, count in sorted(status_counts.items()):
        lines.append(f"| {status} | {count} |")
    lines.extend(["", "| assembly_level | count |", "| --- | ---: |"])
    for assembly_level, count in sorted(assembly_counts.items()):
        lines.append(f"| {assembly_level} | {count} |")
    lines.extend(
        [
            "",
            "## Genome Length Tables",
            "",
            "Each species with a genome FASTA writes an audit genome-length table and a chromosome-only three-column table for circlize.",
            "",
            "| species_id | assembly_level | chromosome_analysis_ready | genome_seq_count | chromosome_seq_count | unassembled_seq_count | organelle_seq_count | chromosome_bp |",
            "| --- | --- | --- | ---: | ---: | ---: | ---: | ---: |",
        ]
    )
    for row in qc_rows:
        lines.append(
            "| {species_id} | {assembly_level} | {chromosome_analysis_ready} | {genome_seq_count} | {chromosome_seq_count} | {unassembled_seq_count} | {organelle_seq_count} | {chromosome_bp} |".format(
                **row
            )
        )
    if missing_chromosome_rows:
        lines.extend(
            [
                "",
                "Species with genome FASTA but no chromosome-classified sequences may need a species-specific chromosome rename or scaffold-to-chromosome map before circlize or JCVI visualization.",
                "",
                "| species_id | genome_seq_count | unassembled_seq_count |",
                "| --- | ---: | ---: |",
            ]
        )
        for row in missing_chromosome_rows:
            lines.append(f"| {row['species_id']} | {row['genome_seq_count']} | {row['unassembled_seq_count']} |")
    lines.append("")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def _split_csv(value: str | None) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in value.split(",") if item.strip()]


def resolve_output_paths(args: argparse.Namespace) -> dict[str, Path]:
    outdir = args.outdir
    return {
        "out_root": args.out_root or outdir / "species_clean_bank",
        "manifest": args.manifest or outdir / "species_clean_bank_manifest.tsv",
        "qc": args.qc or outdir / "species_clean_bank_qc.tsv",
        "qc_excel": args.qc_excel or outdir / "species_clean_bank_qc.xlsx",
        "failed": args.failed or outdir / "species_clean_bank_failed.tsv",
        "summary": args.summary or outdir / "species_clean_bank_summary.md",
        "species_info_txt": args.species_info_txt or outdir / "species_info.txt",
        "species_info_tsv": args.species_info_tsv or outdir / "species_info.tsv",
        "species_tree_dir": args.species_tree_dir or outdir / "species_tree",
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--raw-root", required=True, type=Path)
    parser.add_argument("--outdir", default=Path("results"), type=Path)
    parser.add_argument("--out-root", default=None, type=Path)
    parser.add_argument("--manifest", default=None, type=Path)
    parser.add_argument("--qc", default=None, type=Path)
    parser.add_argument("--qc-excel", default=None, type=Path)
    parser.add_argument("--failed", default=None, type=Path)
    parser.add_argument("--summary", default=None, type=Path)
    parser.add_argument("--species-info-txt", default=None, type=Path)
    parser.add_argument("--species-info-tsv", default=None, type=Path)
    parser.add_argument("--species-tree-dir", default=None, type=Path)
    parser.add_argument("--species-tree-source", choices=["none", "user", "timetree"], default="none")
    parser.add_argument("--species-tree-user-tree", default=None, type=Path)
    parser.add_argument("--species-tree-timetree-run", action="store_true")
    parser.add_argument("--include", default="all")
    parser.add_argument("--exclude", default="")
    parser.add_argument("--require-cds", action="store_true")
    parser.add_argument("--require-genome", action="store_true")
    parser.add_argument("--require-gff3", action="store_true")
    parser.add_argument(
        "--large-file-mode",
        choices=["copy", "symlink", "skip"],
        default="symlink",
        help="How to materialize genome/GFF3 files in raw/ and clean/: symlink by default to save space, copy for portable archives, or skip to reference originals in the manifest.",
    )
    parser.add_argument(
        "--scaffold-chromosome-min-bp",
        type=int,
        default=0,
        help="Promote numbered scaffold IDs such as scaffold_1 to chromosome length rows when their length is at least this many bp; 0 disables this conservative override.",
    )
    args = parser.parse_args()

    paths = resolve_output_paths(args)
    include = "all" if args.include == "all" else _split_csv(args.include)
    manifest_rows, qc_rows, failed_rows = build_species_clean_bank(
        raw_root=args.raw_root,
        out_root=paths["out_root"],
        include=include,
        exclude=_split_csv(args.exclude),
        cds_required=args.require_cds,
        genome_required=args.require_genome,
        gff3_required=args.require_gff3,
        large_file_mode=args.large_file_mode,
        scaffold_chromosome_min_bp=args.scaffold_chromosome_min_bp,
    )
    write_rows(paths["manifest"], MANIFEST_FIELDS, manifest_rows)
    write_rows(paths["qc"], QC_FIELDS, qc_rows)
    write_qc_excel(paths["qc_excel"], qc_rows)
    write_rows(paths["failed"], FAILED_FIELDS, failed_rows)
    write_summary(paths["summary"], qc_rows, failed_rows)
    species_rows = write_species_info(paths["species_info_txt"], paths["species_info_tsv"], qc_rows)
    prepare_species_tree_outputs(
        source=args.species_tree_source,
        user_tree=args.species_tree_user_tree,
        species_rows=species_rows,
        tree_dir=paths["species_tree_dir"],
        timetree_run=args.species_tree_timetree_run,
    )
    print(f"Built species clean bank for {len(qc_rows)} species at {paths['out_root']}")


if __name__ == "__main__":
    main()
