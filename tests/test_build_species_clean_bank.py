import csv
import subprocess
import sys
from pathlib import Path

from bin.genefam.build_species_clean_bank import build_genome_length_rows


REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPT = REPO_ROOT / "bin/genefam/build_species_clean_bank.py"


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def write_demo_species(root: Path, species: str = "Demo_species") -> Path:
    species_dir = root / species
    species_dir.mkdir(parents=True)
    (species_dir / f"{species}.pep.fa").write_text(
        ">GeneA.1|PACid:1\nMAA*\n"
        ">GeneA.2|PACid:2\nMAAAA*\n"
        ">GeneB.1|PACid:3\nMCC*\n",
        encoding="utf-8",
    )
    (species_dir / f"{species}.cds.fa").write_text(
        ">GeneA.1|PACid:1\nATGGCTGCTTAA\n"
        ">GeneA.2|PACid:2\nATGGCTGCTGCTGCTTAA\n"
        ">GeneB.1|PACid:3\nATGTGTTGTTAA\n",
        encoding="utf-8",
    )
    (species_dir / f"{species}.genome.fa").write_text(
        ">chr1 chromosome 1\nATGCATGC\n"
        ">scaffold_1 unplaced scaffold\nATGC\n"
        ">ChrC chloroplast\nATGCAT\n",
        encoding="utf-8",
    )
    (species_dir / f"{species}.gff3").write_text(
        "chr1\tDemo\tgene\t1\t900\t.\t+\t.\tID=GeneA;Name=GeneA\n"
        "chr1\tDemo\tmRNA\t1\t900\t.\t+\t.\tID=PAC:1;Name=GeneA.1;Parent=GeneA\n"
        "chr1\tDemo\tmRNA\t1\t900\t.\t+\t.\tID=PAC:2;Name=GeneA.2;Parent=GeneA\n"
        "chr1\tDemo\tgene\t1000\t1800\t.\t+\t.\tID=GeneB;Name=GeneB\n"
        "chr1\tDemo\tmRNA\t1000\t1800\t.\t+\t.\tID=PAC:3;Name=GeneB.1;Parent=GeneB\n",
        encoding="utf-8",
    )
    return species_dir


def test_build_genome_length_rows_can_promote_large_numbered_scaffolds_to_chromosomes(tmp_path):
    genome = tmp_path / "setaria_like.genome.fa"
    genome.write_text(
        ">scaffold_1\n" + "A" * 20 + "\n"
        ">scaffold_2\n" + "A" * 15 + "\n"
        ">scaffold_10\n" + "A" * 5 + "\n",
        encoding="utf-8",
    )

    default_rows, default_chromosomes = build_genome_length_rows(genome)
    promoted_rows, promoted_chromosomes = build_genome_length_rows(genome, scaffold_chromosome_min_bp=10)

    assert [row["SeqType"] for row in default_rows] == ["unassembled", "unassembled", "unassembled"]
    assert default_chromosomes == []
    assert [row["SeqType"] for row in promoted_rows] == ["chromosome", "chromosome", "unassembled"]
    assert promoted_chromosomes == [
        {"Chr": "scaffold_1", "Start": "1", "End": "20"},
        {"Chr": "scaffold_2", "Start": "1", "End": "15"},
    ]


def test_build_species_clean_bank_writes_raw_clean_audit_and_global_tables(tmp_path):
    raw_root = tmp_path / "species_bank"
    write_demo_species(raw_root)
    out_root = tmp_path / "species_clean_bank"
    manifest = tmp_path / "species_clean_bank_manifest.tsv"
    qc = tmp_path / "species_clean_bank_qc.tsv"
    qc_excel = tmp_path / "species_clean_bank_qc.xlsx"
    failed = tmp_path / "species_clean_bank_failed.tsv"
    summary = tmp_path / "species_clean_bank_summary.md"
    species_info_txt = tmp_path / "species_info.txt"
    species_info_tsv = tmp_path / "species_info.tsv"

    completed = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--raw-root",
            str(raw_root),
            "--out-root",
            str(out_root),
            "--manifest",
            str(manifest),
            "--qc",
            str(qc),
            "--qc-excel",
            str(qc_excel),
            "--failed",
            str(failed),
            "--summary",
            str(summary),
            "--species-info-txt",
            str(species_info_txt),
            "--species-info-tsv",
            str(species_info_tsv),
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode == 0, completed.stderr
    species_root = out_root / "Demo_species"
    assert (species_root / "raw/Demo_species.pep.fa").exists()
    assert (species_root / "raw/Demo_species.cds.fa").exists()
    assert (species_root / "raw/Demo_species.genome.fa").exists()
    assert (species_root / "raw/Demo_species.gff3").exists()
    assert (species_root / "raw/Demo_species.pep.fa").is_symlink()
    assert (species_root / "raw/Demo_species.cds.fa").is_symlink()
    assert (species_root / "raw/Demo_species.genome.fa").is_symlink()
    assert (species_root / "raw/Demo_species.gff3").is_symlink()
    protein_clean = species_root / "clean/Demo_species.protein.clean.fa"
    cds_clean = species_root / "clean/Demo_species.cds.clean.fa"
    assert protein_clean.read_text(encoding="utf-8") == ">GeneA\nMAAAA\n>GeneB\nMCC\n"
    assert cds_clean.read_text(encoding="utf-8") == ">GeneA\nATGGCTGCTGCTGCTTAA\n>GeneB\nATGTGTTGTTAA\n"
    assert (species_root / "clean/Demo_species.genome.fa").exists()
    assert (species_root / "clean/Demo_species.gff3").exists()
    assert not protein_clean.is_symlink()
    assert not cds_clean.is_symlink()
    assert (species_root / "clean/Demo_species.genome.fa").is_symlink()
    assert (species_root / "clean/Demo_species.gff3").is_symlink()
    assert (species_root / "clean/Demo_species.chromosome.lengths.tsv").exists()
    assert (species_root / "audit/Demo_species.gene_id_map.tsv").exists()
    assert (species_root / "audit/Demo_species.id_resolution_rules.tsv").exists()
    assert (species_root / "audit/Demo_species.genome.lengths.tsv").exists()
    assert (species_root / "audit/Demo_species.representative_transcripts.tsv").exists()
    assert (species_root / "audit/Demo_species.preprocess_qc.tsv").exists()
    assert (species_root / "audit/Demo_species.preprocess_warnings.tsv").exists()
    chromosome_rows = read_tsv(species_root / "clean/Demo_species.chromosome.lengths.tsv")
    assert chromosome_rows == [{"Chr": "chr1", "Start": "1", "End": "8"}]
    genome_rows = read_tsv(species_root / "audit/Demo_species.genome.lengths.tsv")
    assert [row["SeqType"] for row in genome_rows] == ["chromosome", "unassembled", "organelle"]

    manifest_rows = read_tsv(manifest)
    assert manifest_rows[0]["species_id"] == "Demo_species"
    assert manifest_rows[0]["status"] == "pass"
    assert manifest_rows[0]["protein"].endswith("clean/Demo_species.protein.clean.fa")
    assert manifest_rows[0]["cds"].endswith("clean/Demo_species.cds.clean.fa")
    assert manifest_rows[0]["genome"].endswith("clean/Demo_species.genome.fa")
    assert manifest_rows[0]["gff3"].endswith("clean/Demo_species.gff3")
    assert manifest_rows[0]["genome_lengths"].endswith("audit/Demo_species.genome.lengths.tsv")
    assert manifest_rows[0]["chromosome_lengths"].endswith("clean/Demo_species.chromosome.lengths.tsv")
    assert manifest_rows[0]["id_resolution_rules"].endswith("audit/Demo_species.id_resolution_rules.tsv")

    qc_rows = read_tsv(qc)
    assert qc_rows[0]["raw_pep_count"] == "3"
    assert qc_rows[0]["clean_pep_count"] == "2"
    assert qc_rows[0]["raw_cds_count"] == "3"
    assert qc_rows[0]["clean_cds_count"] == "2"
    assert qc_rows[0]["genome_seq_count"] == "3"
    assert qc_rows[0]["chromosome_seq_count"] == "1"
    assert qc_rows[0]["unassembled_seq_count"] == "1"
    assert qc_rows[0]["organelle_seq_count"] == "1"
    assert qc_rows[0]["total_genome_bp"] == "18"
    assert qc_rows[0]["chromosome_bp"] == "8"
    assert qc_rows[0]["assembly_level"] == "chromosome"
    assert qc_rows[0]["chromosome_analysis_ready"] == "TRUE"
    assert qc_rows[0]["terminal_stop_removed_count"] == "3"
    assert qc_rows[0]["cds_match_rate"] == "1.0000"
    assert qc_rows[0]["status"] == "pass"
    assert qc_excel.exists()
    from openpyxl import load_workbook

    workbook = load_workbook(qc_excel, read_only=True)
    assert workbook.sheetnames == ["species_qc", "assembly_summary"]
    sheet = workbook["species_qc"]
    assert [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))] == [
        "species_id",
        "raw_pep_count",
        "raw_cds_count",
        "clean_pep_count",
        "clean_cds_count",
        "genome_seq_count",
        "chromosome_seq_count",
        "unassembled_seq_count",
        "organelle_seq_count",
        "total_genome_bp",
        "chromosome_bp",
        "assembly_level",
        "chromosome_analysis_ready",
        "gff3_transcript_map_count",
        "gff3_mapping_rate",
        "fallback_mapping_rate",
        "cds_match_rate",
        "terminal_stop_removed_count",
        "warning_count",
        "status",
        "note",
    ]
    assert read_tsv(failed) == []
    assert "Species Clean Bank Summary" in summary.read_text(encoding="utf-8")
    assert species_info_txt.read_text(encoding="utf-8") == "Demo species\n"
    assert read_tsv(species_info_tsv) == [{"species_id": "Demo_species", "latin_name": "Demo species"}]


def test_build_species_clean_bank_defaults_to_results_outdir(tmp_path):
    raw_root = tmp_path / "species_bank"
    write_demo_species(raw_root)

    completed = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--raw-root",
            str(raw_root),
        ],
        cwd=tmp_path,
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode == 0, completed.stderr
    assert "results/species_clean_bank" in completed.stdout
    assert (tmp_path / "results/species_clean_bank/Demo_species/clean/Demo_species.protein.clean.fa").exists()
    assert (tmp_path / "results/species_clean_bank_manifest.tsv").exists()
    assert (tmp_path / "results/species_clean_bank_qc.tsv").exists()
    assert (tmp_path / "results/species_clean_bank_qc.xlsx").exists()
    assert (tmp_path / "results/species_clean_bank_failed.tsv").exists()
    assert (tmp_path / "results/species_clean_bank_summary.md").exists()
    assert (tmp_path / "results/species_info.txt").read_text(encoding="utf-8") == "Demo species\n"
    assert (tmp_path / "results/species_info.tsv").exists()


def test_build_species_clean_bank_accepts_custom_outdir(tmp_path):
    raw_root = tmp_path / "species_bank"
    outdir = tmp_path / "custom_result"
    write_demo_species(raw_root)

    completed = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--raw-root",
            str(raw_root),
            "--outdir",
            str(outdir),
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode == 0, completed.stderr
    assert (outdir / "species_clean_bank/Demo_species/clean/Demo_species.protein.clean.fa").exists()
    assert (outdir / "species_clean_bank_manifest.tsv").exists()
    assert (outdir / "species_clean_bank_qc.xlsx").exists()


def test_build_species_clean_bank_can_symlink_large_genome_inputs_for_testing(tmp_path):
    raw_root = tmp_path / "species_bank"
    species_dir = write_demo_species(raw_root)
    outdir = tmp_path / "results"

    completed = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--raw-root",
            str(raw_root),
            "--outdir",
            str(outdir),
            "--large-file-mode",
            "symlink",
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode == 0, completed.stderr
    species_root = outdir / "species_clean_bank/Demo_species"
    raw_genome = species_root / "raw/Demo_species.genome.fa"
    raw_gff3 = species_root / "raw/Demo_species.gff3"
    clean_genome = species_root / "clean/Demo_species.genome.fa"
    clean_gff3 = species_root / "clean/Demo_species.gff3"
    assert raw_genome.is_symlink()
    assert raw_gff3.is_symlink()
    assert clean_genome.is_symlink()
    assert clean_gff3.is_symlink()
    assert raw_genome.resolve() == (species_dir / "Demo_species.genome.fa").resolve()
    assert raw_gff3.resolve() == (species_dir / "Demo_species.gff3").resolve()
    assert clean_genome.resolve() == (species_dir / "Demo_species.genome.fa").resolve()
    assert clean_gff3.resolve() == (species_dir / "Demo_species.gff3").resolve()
    manifest_rows = read_tsv(outdir / "species_clean_bank_manifest.tsv")
    assert manifest_rows[0]["genome"].endswith("clean/Demo_species.genome.fa")
    assert manifest_rows[0]["gff3"].endswith("clean/Demo_species.gff3")
    assert read_tsv(species_root / "clean/Demo_species.chromosome.lengths.tsv") == [
        {"Chr": "chr1", "Start": "1", "End": "8"}
    ]


def test_build_species_clean_bank_records_incomplete_species_without_stopping(tmp_path):
    raw_root = tmp_path / "species_bank"
    write_demo_species(raw_root)
    incomplete = raw_root / "Incomplete_species"
    incomplete.mkdir(parents=True)
    (incomplete / "all.pep").write_text(">GeneX\nMAA\n", encoding="utf-8")
    outdir = tmp_path / "results"

    completed = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--raw-root",
            str(raw_root),
            "--outdir",
            str(outdir),
            "--require-cds",
            "--require-genome",
            "--require-gff3",
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode == 0, completed.stderr
    assert (outdir / "species_clean_bank/Demo_species/clean/Demo_species.protein.clean.fa").exists()
    qc_rows = {row["species_id"]: row for row in read_tsv(outdir / "species_clean_bank_qc.tsv")}
    assert qc_rows["Demo_species"]["status"] == "pass"
    assert qc_rows["Incomplete_species"]["status"] == "missing_required_input"
    assert qc_rows["Incomplete_species"]["note"] == "missing required input file(s): pep, cds, genome, gff3"
    failed_rows = read_tsv(outdir / "species_clean_bank_failed.tsv")
    assert failed_rows == [
        {
            "species_id": "Incomplete_species",
            "status": "missing_required_input",
            "reason": "missing required input file(s): pep, cds, genome, gff3",
        }
    ]


def test_build_species_clean_bank_can_copy_user_species_tree(tmp_path):
    raw_root = tmp_path / "species_bank"
    write_demo_species(raw_root)
    user_tree = tmp_path / "species_tree.nwk"
    user_tree.write_text("(Demo species:1);\n", encoding="utf-8")
    outdir = tmp_path / "results"

    completed = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--raw-root",
            str(raw_root),
            "--outdir",
            str(outdir),
            "--species-tree-source",
            "user",
            "--species-tree-user-tree",
            str(user_tree),
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode == 0, completed.stderr
    assert (outdir / "species_tree/species_tree.nwk").read_text(encoding="utf-8") == "(Demo species:1);\n"
    assert read_tsv(outdir / "species_tree/species_tree_status.tsv") == [
        {
            "source": "user",
            "status": "available",
            "species_count": "1",
            "tree": str((outdir / "species_tree/species_tree.nwk").resolve()),
            "note": "Copied user-provided Newick tree",
        }
    ]


def test_build_species_clean_bank_disables_species_tree_and_removes_stale_managed_outputs(tmp_path):
    raw_root = tmp_path / "species_bank"
    write_demo_species(raw_root)
    outdir = tmp_path / "results"
    stale_dir = outdir / "species_tree"
    stale_dir.mkdir(parents=True)
    (stale_dir / "species_tree.nwk").write_text("(Old_species:1);\n", encoding="utf-8")
    (stale_dir / "timetree_species_input.txt").write_text("Old species\n", encoding="utf-8")
    (stale_dir / "timetree_species_validation.tsv").write_text("old\n", encoding="utf-8")

    completed = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--raw-root",
            str(raw_root),
            "--outdir",
            str(outdir),
            "--species-tree-source",
            "none",
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode == 0, completed.stderr
    assert not (stale_dir / "species_tree.nwk").exists()
    assert not (stale_dir / "timetree_species_input.txt").exists()
    assert not (stale_dir / "timetree_species_validation.tsv").exists()
    assert read_tsv(stale_dir / "species_tree_status.tsv") == [
        {
            "source": "none",
            "status": "disabled",
            "species_count": "1",
            "tree": "",
            "note": "No species tree configured; downstream species-tree panels should be skipped.",
        }
    ]


def test_build_species_clean_bank_records_missing_user_species_tree_without_stopping(tmp_path):
    raw_root = tmp_path / "species_bank"
    write_demo_species(raw_root)
    outdir = tmp_path / "results"
    missing_tree = tmp_path / "missing_species_tree.nwk"

    completed = subprocess.run(
        [
            sys.executable,
            str(SCRIPT),
            "--raw-root",
            str(raw_root),
            "--outdir",
            str(outdir),
            "--species-tree-source",
            "user",
            "--species-tree-user-tree",
            str(missing_tree),
        ],
        check=False,
        capture_output=True,
        text=True,
    )

    assert completed.returncode == 0, completed.stderr
    status = read_tsv(outdir / "species_tree/species_tree_status.tsv")
    assert status == [
        {
            "source": "user",
            "status": "missing_input",
            "species_count": "1",
            "tree": "",
            "note": "User species tree was not provided or does not exist; downstream species-tree panels should be skipped.",
        }
    ]
